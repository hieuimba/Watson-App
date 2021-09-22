import streamlit as st
import pandas as pd
import time
import datetime
from datetime import datetime as dt
from ta import volatility
import numpy as np
import streamlit.components.v1 as components # for html

import os

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

import mysql.connector
from sqlalchemy import create_engine
import sqlalchemy


##-------------------------------------------------SETTINGS-----------------------------------------------------------##
# layout setting
st.set_page_config(layout='wide', page_title = 'Alfred 4', page_icon = '📈')
initial_risk = 20 #<--------using static risk
hide_menu_style = """
       <style>
       #MainMenu {visibility: hidden; }
       footer {visibility: hidden;}
       </style>
       """
st.markdown(hide_menu_style, unsafe_allow_html=True)

# mysql database set-up
user = 'source'
password = r'Source:24.76.105.42'
host = '3.99.99.227'

@st.cache(hash_funcs={sqlalchemy.engine.base.Engine: id})
def db_connect(db):
    return create_engine(f"mysql://{user}:{password}@{host}/{db}")

@st.cache(allow_output_mutation=True, hash_funcs={sqlalchemy.engine.base.Engine: id})
def get_table(connection, table, index = None):
    return pd.read_sql_table(table, con = connection, index_col = index)

@st.cache(allow_output_mutation=True, hash_funcs={sqlalchemy.engine.base.Engine: id})
def run_query(connection, query):
    return pd.read_sql_query(query, con = connection)

positions = db_connect('positions')
prices = db_connect('prices')
# export_log = pd.read_csv(r'C:\Users\hieui\Desktop\Data\dashboard_log\export_log.csv')


# make database connections
open_positions = get_table(positions, 'open_positions', 'symbol')
open_orders = get_table(positions, 'open_orders', 'symbol')
closed_orders = get_table(positions, 'closed_orders', 'symbol')
closed_positions = closed_orders.copy()
etf = get_table(prices, 'etf_price', None)
updated = run_query(prices, "SELECT Updated FROM symbol_list LIMIT 1")

##---------------------------------------------DASHBOARD ELEMENTS-----------------------------------------------------##

##----------LAYOUT SETUP----------
option = st.radio('', options = ['Positions','Position Calc','Orders','Sectors','Scanner','Watchlist'])
st.write('<style>div.row-widget.stRadio > div{flex-direction:row;}</style>', unsafe_allow_html = True)

st.caption(f'Updated: {updated.iat[0,0]}')
st.markdown(f"<h1 style='text-align: center; color: black;'>{option}</h1>", unsafe_allow_html=True)

##----------POSITIONS SCREEN------
if option == 'Positions':
    # header-row calcs
    unrealized_pnl = '{0:.2f}'.format(open_positions['unrlzd p&l'].sum() / initial_risk) + ' R'
    realized_pnl = '{0:.2f}'.format(closed_orders['rlzd p&l'].sum() / initial_risk) + ' R'
    total_pnl = '{0:.2f}'.format((open_positions['unrlzd p&l'].sum() + closed_orders['rlzd p&l'].sum()) / initial_risk,
                                 2) + ' R'
    atr_risk = '{0:.2f}'.format(open_positions['atr risk'].sum() / initial_risk) + ' R'
    total_open_risk = '{0:.2f}'.format(open_positions['open risk'].sum() / initial_risk) + ' R'

    # prepare open positions table
    open_positions = open_positions.drop(columns = ['atr', 'atr risk'])
    open_positions['rlzd p&l'] /= initial_risk
    open_positions['unrlzd p&l'] /= initial_risk
    open_positions['open risk'] /= initial_risk

    # prepare closed positions table
    ls = []
    for i in range(0, len(closed_positions), 1):
        if closed_positions.iloc[i]['action'] == 'BUY' and closed_positions.iloc[i]['rlzd p&l'] != 0:
            ls.append('short')
        elif closed_positions.iloc[i]['action'] == 'SELL' and closed_positions.iloc[i]['rlzd p&l'] != 0:
            ls.append('long')
        else:
            ls.append(np.nan)
    closed_positions['l/s'] = ls
    closed_positions = closed_positions[closed_positions['status'] == 'Filled']  # filter for filled trades only
    closed_positions = closed_positions[(closed_positions['l/s'] == 'long') | (closed_positions['l/s'] == 'short')]
    closed_positions = closed_positions.drop(
        columns = ['action', 'type', 'status', 'time', 'price', 'filled qty', 'comm.'])
    closed_positions['entry'], closed_positions['stop'], closed_positions['target'], closed_positions[
        'unrlzd p&l'] = 0, 0, 0, 0
    closed_positions.rename(columns = {'filled at': 'exit'}, inplace = True)
    closed_positions = closed_positions[['l/s', 'qty', 'entry', 'exit', 'stop', 'target', 'unrlzd p&l', 'rlzd p&l']]
    closed_positions['rlzd p&l'] /= initial_risk
    closed_positions['unrlzd p&l'] /= initial_risk

    '---'
    one, two, three, four = st.columns([2,4.5,1.5,2])
    # with four:
    #     risk_parameters = st.radio('',
    #                               ('ATR', 'cVaR 95%', 'cVaR 99%'))
    with two:
        st.header(f'Current P&L: {total_pnl}')
        st.text(f'Unrealized: {unrealized_pnl}')
        st.text(f'Realized: {realized_pnl}')
    with three:
        st.header('Risk:')
        st.text(f'Total open risk: {total_open_risk}')
        # if risk_parameters == 'ATR':
        #     st.text(f'Daily risk: {atr_risk}')
        # elif risk_parameters == 'cVaR 95%':
        #     st.text('Daily risk: {cVaR 95%}')
        # elif risk_parameters == 'cVaR 99%':
        #     st.text('Daily risk: {cVaR 99%}')
    # open positions
    '---'
    one, two, three = st.columns([2, 6, 2])
    with two:
        open_positions.sort_values(by= 'unrlzd p&l', inplace = True, ascending = False)
        st.text(f'{len(open_positions)} trade in progress...' if len(open_positions) == 1 else
                f'{len(open_positions)} trades in progress...')
        st.table(open_positions.style.format({'qty': '{0:.2f}',
                                                     'entry': '{0:.2f}',
                                                     'last': '{0:.2f}',
                                                     'stop': '{0:.2f}',
                                                     'target': '{0:.2f}',
                                                     'unrlzd p&l': '{0:.2f} R',
                                                     'rlzd p&l': '{0:.2f} R',
                                                     'open risk': '{0:.2f} R'},
                                                      na_rep = 'N/A'))
        if len(closed_positions) > 0:
            closed_positions.sort_values(by = 'rlzd p&l', inplace = True, ascending = False)
            st.text(f'{len(closed_positions)} closed trade' if len(closed_positions) == 1 else
                    f'{len(closed_positions)} closed trades')
            st.table(closed_positions.style.format({'qty': '{0:.2f}',
                                                           'entry': '{0:.2f}',
                                                           'exit': '{0:.2f}',
                                                           'stop': '{0:.2f}',
                                                           'target': '{0:.2f}',
                                                           'unrlzd p&l': '{0:.2f} R',
                                                           'rlzd p&l': '{0:.2f} R'},
                                                            na_rep = 'N/A'))
    # selection box & tradingview widget
    one, two, three = st.columns([2, 6, 2])
    with two:
        selection_df = pd.DataFrame()
        spy = []
        spy.append('SPY')
        selection_df['list'] = list(open_positions.index.values) + list(closed_positions.index.values) + spy
        HtmlFile = open(r'C:\Users\hieui\Desktop\Python\trade\venv_01\Code\tradingview.html', 'r', encoding = 'utf-8')
        source_code = HtmlFile.read()

        selection = st.selectbox('', (selection_df))
        source_code = source_code.replace("AAPL", selection)
        components.html(source_code, height = 800)

##----------ORDERS SCREEN---------
if option == 'Orders':
    '---'
    st.text(f'{len(open_orders)} open orders')
    st.dataframe(open_orders.style.format({'qty': '{0:.2f}',
                                              'price': '{0:.2f}'},
                                              na_rep = 'N/A'))
    st.text(f'{len(closed_orders)} closed orders')
    st.dataframe(closed_orders.style.format({'qty': '{0:.2f}',
                                                'price': '{0:.2f}',
                                                'filled at':  '{0:.2f}',
                                                'filled qty': '{0:.2f}',
                                                'comm.': '{0:.2f}',
                                                'rlzd p&l': '{0:.2f}'},
                                                na_rep = 'N/A'))
    export = st.button('Export to Journal')
    # if export:
    #     begin_time = datetime.datetime.now()
    #     journal = r'G:\My Drive\journal_final.xlsx'
    #     wb = load_workbook(filename = journal)
    #     ws = wb['import']
    #     for row in ws['A1:Z100']:
    #         for cell in row:
    #             cell.value = None
    #     rows = dataframe_to_rows(export_log, index = False)
    #     st.write(export_log)
    #     for r_idx, row in enumerate(rows, 1):
    #         for c_idx, value in enumerate(row, 1):
    #             ws.cell(row = r_idx, column = c_idx, value = value)
    #     for row in ws['A1:Z100']:
    #         for cell in row:
    #             cell.font = Font(name='Calibri', size=11)
    #     wb.save(journal)
    #     wb.close()
    #     st.success('Done!')
    #     delta = str(datetime.timedelta(seconds = (datetime.datetime.now() - begin_time).total_seconds()))
    #     st.text(f'Total time:  {delta}')

##----------SECTORS SCREEN--------
sector_list = ['SPY', 'XLE', 'XLI', 'XLK', 'XLY', 'XLF', 'XLB', 'XLP', 'XLV', 'XLU', 'XLRE', 'XLC', 'IWM', 'QQQ']
if option == 'Sectors':
    '---'
    beta_list = []
    matrix = pd.DataFrame()
    std_dev = pd.DataFrame()

    one, two = st.columns([1,5])
    with one:
        options = st.radio('Select period:', options = ['1 M', '3 M', '6 M', '1 Y', '2 Y'], help ='1 month = 21 trading days')
        if options == '1 M':
            period = 21
        elif options == '3 M':
            period = 63
        elif options == '6 M':
            period = 126
        elif options == '1 Y':
            period = 252
        elif options == '2 Y':
            period = 504

    with two:
        sector_name = ['S&P 500 ETF', 'Energy', 'Industrials', 'Technology', 'Consumer Discretionary',
                       'Financials', 'Materials', 'Consumer Staples', 'Health Care', 'Utilities',
                       'Real Estate', 'Communication Services', 'Russell 2000 ETF', 'Invesco QQQ Trust']
        sector_trans = pd.DataFrame(data = sector_name, columns = ['Name'], index = sector_list)
        st.table(sector_trans.T)

    spy = run_query(prices, "SELECT * FROM etf_price WHERE symbol = 'SPY'")
    spy['return%'] = spy['Close'].pct_change(1) * 100
    spy = spy.tail(period)
    spy['var'] = spy['return%'].var()
    for i in range(0, len(sector_list)):
        sector = etf[etf['Symbol'] == sector_list[i]].copy()
        sector['return%'] = sector['Close'].pct_change(1) * 100
        sector = sector.tail(period)
        matrix[f'{sector_list[i]}'] = sector['return%'].values
        corr_matrix = matrix.corr()

        sector['bm_return%'] = spy['return%'].to_list()
        cov_df = sector[['return%', "bm_return%"]].cov()
        bm_var = spy.iloc[-1]['var']
        beta = cov_df.iloc[1, 0] / bm_var
        beta_list.append(beta)

    temp = pd.DataFrame(index = sector_list)
    temp['Beta'] = beta_list
    temp = temp.T
    corr_matrix = corr_matrix.append(temp)
    u = corr_matrix.index.get_level_values(0)
    corr_matrix = corr_matrix.style.background_gradient(cmap = 'Oranges', axis = None, low = -0.5, subset = pd.IndexSlice[u[:-1], :])
    corr_matrix = corr_matrix.background_gradient(cmap = 'Greens', axis = None, subset = pd.IndexSlice[u[-1], :])
    st.table(corr_matrix.format(precision = 3))

##----------CALC SCREEN-----------
if option == 'Position Calc':
    '---'
    # etf = get_table(prices, 'etf_price', None)
    # stock = get_table(prices, 'stock_price', None)
    # stock_list = prices.execute("""SELECT DISTINCT Symbol
    #                             FROM stock_price;""")
    # symbol = 'AMD'
    # ### stock = prices.execute(f"SELECT * FROM stock_price WHERE symbol = '{symbol}' LIMIT 20")
    # stock = pd.read_sql_query(f"SELECT * FROM stock_price WHERE symbol = '{symbol}'", con=prices)
    # st.write(stock)
    # select * from stock_price where symbol = 'AAPL';
    # st.write(stock.fetchall())
    # stock_list =
    # st.write(stock_list.fetchall())
    # stock_path = r'C:\UA\Data\All Top US Exchange Stocks by Market Cap'
    # stock_path = os.path.join(stock_path, '')
    # stock_files = os.listdir(stock_path)

    # etf_path = r'C:\UA\Data\ETF'
    # etf_path = os.path.join(etf_path, '')
    # etf_files = os.listdir(etf_path)
    symbols = run_query(prices, "SELECT symbol FROM symbol_list")
    symbols = symbols['symbol'].to_list()
    # for stock in stock_files:
    #     symbols.append(stock.split('.csv')[0])
    # for etf in etf_files:
    #     symbols.append(etf.split('.csv')[0])

    beta_list = []
    matrix = pd.DataFrame()
    std_dev = pd.DataFrame()

    one, two, three = st.columns([2,3,1])
    with one:
        risk_options = [10, 20, 30, 50, 80, 100]
        risk = st.selectbox(label = '$ Risk', options = risk_options, index = 1)
        entry = st.number_input(label='Entry', value = 2.00, step = 0.1)
        stop = st.number_input(label='Stop', value = 1.00, step = 0.1)
        target = entry + (entry - stop)
        distance = round(entry - stop, 2)
        distance_percent = round(abs(distance) / entry, 2) * 100
        size = round(risk / abs(distance), 3)
        dollar_size = round(size * entry, 2)
        direction = 'Long' if distance > 0 else 'Short'
        st.number_input('Target', min_value = target, max_value = target, value =target)
        st.text(f'Direction: {direction}')
        st.subheader(f'Size: {size} share' if size == 1 else f'Size: {size} shares')
        ''
        st.text(f'$ Equivalent: $ {dollar_size}')
        st.subheader('R table:')
        price_range = [stop, stop + distance/2, entry, entry+ distance/2, target, entry + distance*2]
        price_range = pd.DataFrame(price_range, columns = ['Price'],
                                                index = ['-1 R', '-1.5 R', '0 R', '0.5 R', '1 R', '2 R']).T
        st.table(price_range.style.format(precision = 2))
    with three:
        options = st.radio('Select period:', options = ['1 M', '3 M', '6 M', '1 Y'], help ='1 month = 21 trading days')
        if options == '1 M':
            period = 21
        elif options == '3 M':
            period = 63
        elif options == '6 M':
            period = 126
        elif options == '1 Y':
            period = 252
    with two:
        symbol = st.multiselect('Select symbols:', options = symbols, default = ['SPY'] + open_positions.index.values.tolist())

        spy = run_query(prices, "SELECT * FROM etf_price WHERE symbol = 'SPY'")
        spy['return%'] = spy['Close'].pct_change(1) * 100
        spy = spy.tail(period)
        spy['var'] = spy['return%'].var()

        for i in range(0, len(symbol)):
            if symbol[i] in sector_list:
                bars = run_query(prices, f"SELECT * FROM etf_price WHERE symbol = '{symbol[i]}'")
                bars = bars.fillna('N/A')
            else:
                bars = run_query(prices, f"SELECT * FROM stock_price WHERE symbol = '{symbol[i]}'")
                bars = bars.fillna('N/A')
            bars['atr'] = volatility.AverageTrueRange(bars['High'], bars['Low'], bars['Close'],
                                                      window = 21).average_true_range()
            bars['return%'] = bars['Close'].pct_change(1) * 100
            if i == len(symbol) -1:
                bars['std dev'] = bars['return%'].rolling(21).std()
                bars['ATR'] = volatility.AverageTrueRange(bars['High'], bars['Low'], bars['Close'],
                                                          window = 21).average_true_range()
                bars['avg vol'] = bars['Volume'].rolling(21).mean()
            bars = bars.tail(period)
            matrix[f'{symbol[i]}'] = bars['return%'].values
            corr_matrix = matrix.corr()

            bars['bm_return%'] = spy['return%'].to_list()
            cov_df = bars[['return%', "bm_return%"]].cov()
            bm_var = spy.iloc[-1]['var']
            beta = cov_df.iloc[1, 0] / bm_var
            beta_list.append(beta)

        temp = pd.DataFrame(index = symbol)
        temp['Beta-M'] = beta_list
        temp = temp.T
        temp2 = pd.DataFrame(index = symbol)
        temp2['Beta-S'] = 0
        temp2 = temp2.T
        corr_matrix = corr_matrix.append(temp)
        corr_matrix = corr_matrix.append(temp2)
        u = corr_matrix.index.get_level_values(0)
        corr_matrix = corr_matrix.style.background_gradient(cmap = 'RdYlGn_r', axis = None,
                                                            subset = pd.IndexSlice[u[:-2], :])
        # corr_matrix = corr_matrix.background_gradient(cmap = 'White', axis = None, subset = pd.IndexSlice[u[-1], :])
        st.table(corr_matrix.format(precision = 3))


        st.text(f"Symbol: {bars.iloc[-1]['Symbol']},    Name: {bars.iloc[-1]['Name']},    Sector: {bars.iloc[-1]['Sector']}")
        st.text(f"Last: {bars.iloc[-1]['Close']},    Relative Volume: {round(bars.iloc[-1]['Volume']/bars.iloc[-1]['avg vol'], 2)}")
        atr = bars.iloc[-1]['ATR']
        st.text(f"Distance: {abs(distance)},    ATR: {round(atr, 2)},    Stop/ATR: {round(distance/atr,  2)}")
        st.text(f"Distance %: {distance_percent} %,   1 Sigma: {round(bars.iloc[-1]['std dev'], 2)} %,    Stop/Sigma: {round(distance_percent/bars.iloc[-1]['std dev'], 2)}")

        ''
        st.subheader('Clipboard:')
        clipboard_dir = r'C:\Users\hieui\Desktop\Data\dashboard_log\clipboard_log.csv'
        clipboard_log = pd.read_csv(clipboard_dir, index_col = 'symbol')
        st.table(clipboard_log.style.format(precision = 2))
    with three:
        # for i in range(0, len(symbol)):
        #     if symbol[i] in sector_list:
        #         bars = pd.read_csv(etf_path + symbol[i] + '.csv').fillna('N/A')
        #     else:
        #         bars = pd.read_csv(stock_path + symbol[i] + '.csv').fillna('N/A')
        #     st.text(f"{bars.iloc[-1]['MarketName']},  {bars.iloc[-1]['Sector']}")
        clipboard_list = []
        st.title('')
        st.title('')
        st.title('')
        st.title('')
        st.title('')
        st.title('')
        st.title('')
        st.title('')
        st.title('')
        st.title('')
        send_order = st.button('Send to IB')
        clipboard_save = st.button('Append to Clipboard')
        clipboard_delete = st.button('Delete last Row')
        clipboard_delete_all = st.button('Delete all')
        if clipboard_save:
            clipboard_list.append([direction, entry, stop, target, size])
            clipboard_df = pd.DataFrame(clipboard_list, columns = ['l/s', 'entry', 'stop', 'target', 'size'], index = [symbol[-1].upper()])
            clipboard_df.to_csv(clipboard_dir, mode = 'a', header = None)
            st.experimental_rerun()
        if clipboard_delete:
            clipboard_log = clipboard_log[:-1]
            clipboard_log.to_csv(clipboard_dir)
            st.experimental_rerun()
        if clipboard_delete_all:
            clipboard_log = clipboard_log[0:0]
            clipboard_log.to_csv(clipboard_dir)
            st.experimental_rerun()
        if send_order:
            # ib_send_order.send(direction, symbol, entry, stop, target, size)
            st.warning('Not available yet')
            st.experimental_rerun()
